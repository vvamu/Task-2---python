#py -m pip install PyPDF2
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.styles.colors import Color, RgbColor
from openpyxl.utils import get_column_letter
import os
import asyncio

async def delete_sheets_containing_string(file_path, resultFileName):
    wb = openpyxl.load_workbook(file_path)

    sheets_to_delete = [sheet for sheet in wb.sheetnames if resultFileName in sheet]

    for sheet_name in sheets_to_delete:
        sheet = wb[sheet_name]
        wb.remove(sheet)

    wb.save(file_path)


async def excel_union_files_to_one_file(directory_path, resultFileName):
    resultFile = os.path.join(directory_path,  resultFileName + "_cont" + ".xlsx")
    resultFile2 = os.path.join(directory_path,  resultFileName + ".xlsx")

    wb = openpyxl.Workbook()
    filenames = os.listdir(directory_path)

    for filename in filenames:
        if(filename.__contains__(resultFileName))  :
            continue
        file_path = os.path.join(directory_path, filename).replace("\\", "/")
        
        try:
            excel_file = openpyxl.load_workbook(file_path)

            for sheet in excel_file.worksheets:
                sheet_name = sheet.title
                new_sheet_title = os.path.splitext(filename)[0]
                 
                
                if sheet_name == "ДО":
                    continue
                
                if sheet_name != "Лист1":
                    new_sheet_title = f"{os.path.splitext(filename)[0]}_{sheet_name}"
                else:
                    new_sheet_title = os.path.splitext(filename)[0]

                ws = wb.create_sheet(title=new_sheet_title)

                skip_rows = 0
                skip_keywords = ["график образовательного", "количество часов учебных занятий"]

                for row in sheet.iter_rows(values_only=True):
                    if skip_rows > 0:
                        skip_rows -= 1
                        continue

                    should_skip_row = False
                    for value in row:
                        if any(keyword.lower() in str(value).lower() for keyword in skip_keywords):
                            if "график образовательного" in str(value).lower():
                                skip_rows = 12
                            elif "количество часов учебных занятий" in str(value).lower():
                                should_skip_row = True
                                break

                        if should_skip_row:
                            break

                        new_row = [value for value in row]
                        ws.append(new_row)
                
                # Copy cell styles and formatting
                for row in ws.iter_rows():
                    for cell in row:
                        source_cell = sheet.cell(row=cell.row, column=cell.column)            
                        new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        
                        new_cell.font = Font(size=6,
                                            bold=source_cell.font.bold)

                        #if isinstance(source_cell.fill.start_color.rgb, str):
                        #    rgb_color = source_cell.fill.start_color.rgb
                        #    if len(rgb_color) == 6:
                        #        r, g, b = rgb_color[:2], rgb_color[2:4], rgb_color[4:]
                        #        my_red = Color(rgb=f'FF{r}{g}{b}')
                        #        new_cell.fill = PatternFill(start_color=my_red, end_color=my_red, fill_type=source_cell.fill.fill_type)

                        if source_cell.fill.start_color.rgb:
                            new_cell.fill = PatternFill(start_color=source_cell.fill.start_color,
                                                        end_color=source_cell.fill.end_color,
                                                        fill_type=source_cell.fill.fill_type)

                        new_cell.border = Border(left=source_cell.border.left, 
                                                right=source_cell.border.right,
                                                top=source_cell.border.top,
                                                bottom=source_cell.border.bottom)
                        
                        new_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                                    vertical=source_cell.alignment.vertical)
                        
                        new_cell.number_format = source_cell.number_format

            wb.save(resultFileName + ".xlsx")
            
        except Exception as e:
            print(f"Error processing file: {resultFileName} - {e}")

        #try:
        #    default_sheet = wb['Sheet']
        #    wb.remove(default_sheet)
        #except Exception as e:
        #    print(f"Error processing default_sheet: {e}")
    print("file already executed 1  " + filename )
    await delete_sheets_containing_string(resultFileName + ".xlsx","output")
    print("file already executed 2 " + filename )





async def main():
    directory_path_1 = "D:/work/Univer/Task 2 - python/Учебные планы — копия/С доп. вып._Учебные планы БАК ФЗО 2023"
    directory_path_2 = "D:/work/Univer/Task 2 - python/Учебные планы — копия/С доп. вып._Учебные планы МАГ ФЗО 2023"

    task1 = asyncio.create_task(excel_union_files_to_one_file(directory_path_1, "С доп. вып._Учебные планы БАК ФЗО 2023"))
    task2 = asyncio.create_task(excel_union_files_to_one_file(directory_path_2, "С доп. вып._Учебные планы МАГ ФЗО 2023"))

    await task1
    await task2
    subprocess.Popen(["powershell.exe", "-File", "notify.ps1"])

asyncio.run(main())